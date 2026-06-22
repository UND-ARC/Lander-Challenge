"""
sensor_logger.py
================
Collects data from:
  - BNO055 #1  (I2C addr 0x28 — default)
  - BNO055 #2  (I2C addr 0x29 — ADDR pin pulled HIGH)
  - MPL3115A2  (I2C addr 0x60 — barometric altimeter)
  - TeraRanger Evo 60m (I2C addr 0x31)

Bus: Raspberry Pi I2C-1  →  Pin 3 = SDA, Pin 5 = SCL

Logs for LOG_DURATION_S seconds (default 30 min) at ~SAMPLE_HZ Hz,
then writes a formatted .xlsx workbook for Kalman filter drift analysis.

Dependencies:
    pip install smbus2 openpyxl
"""

import time
import struct
import logging
from datetime import datetime
from dataclasses import dataclass, field, fields
from typing import Optional

import smbus2
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────────────
I2C_BUS          = 1          # /dev/i2c-1  →  Pin 3 (SDA), Pin 5 (SCL)
LOG_DURATION_S   = 30 * 60   # 30 minutes
SAMPLE_HZ        = 10         # target sample rate
SAMPLE_PERIOD    = 1.0 / SAMPLE_HZ

OUTPUT_FILENAME  = f"sensor_drift_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# I2C Addresses
# ─────────────────────────────────────────────────────────────────────────────
BNO055_ADDR_A   = 0x28   # ADDR pin LOW  (default)
BNO055_ADDR_B   = 0x29   # ADDR pin HIGH
MPL3115_ADDR    = 0x60
TERARANGER_ADDR = 0x31


# ─────────────────────────────────────────────────────────────────────────────
# Data Record
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class SensorRecord:
    timestamp_s:      float = 0.0   # seconds since start

    # BNO055 #1
    bno1_ok:          bool  = False
    bno1_sys_cal:     int   = 0
    bno1_gyro_cal:    int   = 0
    bno1_accel_cal:   int   = 0
    bno1_mag_cal:     int   = 0
    bno1_euler_h:     float = 0.0   # heading  deg
    bno1_euler_r:     float = 0.0   # roll     deg
    bno1_euler_p:     float = 0.0   # pitch    deg
    bno1_gyro_x:      float = 0.0   # rad/s
    bno1_gyro_y:      float = 0.0
    bno1_gyro_z:      float = 0.0
    bno1_accel_x:     float = 0.0   # m/s²
    bno1_accel_y:     float = 0.0
    bno1_accel_z:     float = 0.0
    bno1_linaccel_x:  float = 0.0   # m/s² (gravity removed)
    bno1_linaccel_y:  float = 0.0
    bno1_linaccel_z:  float = 0.0
    bno1_temp_c:      float = 0.0

    # BNO055 #2
    bno2_ok:          bool  = False
    bno2_sys_cal:     int   = 0
    bno2_gyro_cal:    int   = 0
    bno2_accel_cal:   int   = 0
    bno2_mag_cal:     int   = 0
    bno2_euler_h:     float = 0.0
    bno2_euler_r:     float = 0.0
    bno2_euler_p:     float = 0.0
    bno2_gyro_x:      float = 0.0
    bno2_gyro_y:      float = 0.0
    bno2_gyro_z:      float = 0.0
    bno2_accel_x:     float = 0.0
    bno2_accel_y:     float = 0.0
    bno2_accel_z:     float = 0.0
    bno2_linaccel_x:  float = 0.0
    bno2_linaccel_y:  float = 0.0
    bno2_linaccel_z:  float = 0.0
    bno2_temp_c:      float = 0.0

    # MPL3115A2
    mpl_ok:           bool  = False
    mpl_altitude_m:   float = 0.0
    mpl_pressure_pa:  float = 0.0
    mpl_temp_c:       float = 0.0

    # TeraRanger Evo
    evo_ok:           bool  = False
    evo_distance_m:   float = 0.0


# ─────────────────────────────────────────────────────────────────────────────
# BNO055 Driver (bare-metal I2C)
# ─────────────────────────────────────────────────────────────────────────────
class BNO055:
    # Register map (subset)
    REG_CHIP_ID    = 0x00
    REG_OPR_MODE   = 0x3D
    REG_PWR_MODE   = 0x3E
    REG_SYS_TRIGGER= 0x3F
    REG_UNIT_SEL   = 0x3B
    REG_CALIB_STAT = 0x35
    REG_EULER_H_L  = 0x1A   # Euler heading LSB
    REG_GYRO_X_L   = 0x14
    REG_ACCEL_X_L  = 0x08
    REG_LIA_X_L    = 0x28   # Linear acceleration
    REG_TEMP       = 0x34

    OPR_MODE_NDOF  = 0x0C
    OPR_MODE_CONFIG= 0x00
    BNO055_CHIP_ID = 0xA0

    def __init__(self, bus: smbus2.SMBus, addr: int, label: str):
        self.bus   = bus
        self.addr  = addr
        self.label = label
        self.ok    = False
        self._init()

    def _write(self, reg: int, val: int):
        self.bus.write_byte_data(self.addr, reg, val)

    def _read(self, reg: int, length: int) -> bytes:
        return bytes(self.bus.read_i2c_block_data(self.addr, reg, length))

    def _init(self):
        try:
            chip_id = self._read(self.REG_CHIP_ID, 1)[0]
            if chip_id != self.BNO055_CHIP_ID:
                log.error("%s: unexpected chip ID 0x%02X", self.label, chip_id)
                return

            # Reset
            self._write(self.REG_OPR_MODE, self.OPR_MODE_CONFIG)
            time.sleep(0.025)
            self._write(self.REG_SYS_TRIGGER, 0x20)
            time.sleep(0.65)

            # Normal power mode
            self._write(self.REG_PWR_MODE, 0x00)
            time.sleep(0.01)
            self._write(self.REG_SYS_TRIGGER, 0x00)
            time.sleep(0.01)

            # Units: m/s² for accel, rad/s for gyro, degrees for Euler, °C
            self._write(self.REG_UNIT_SEL, 0b00000010)   # bit1 = rad/s gyro
            time.sleep(0.01)

            # NDOF fusion mode
            self._write(self.REG_OPR_MODE, self.OPR_MODE_NDOF)
            time.sleep(0.02)

            self.ok = True
            log.info("%s: initialized at 0x%02X", self.label, self.addr)
        except Exception as e:
            log.error("%s: init failed — %s", self.label, e)

    def _read_s16(self, data: bytes, offset: int) -> int:
        val = (data[offset + 1] << 8) | data[offset]
        return val - 65536 if val > 32767 else val

    def read(self) -> dict:
        out = {k: 0.0 for k in [
            "sys_cal","gyro_cal","accel_cal","mag_cal",
            "euler_h","euler_r","euler_p",
            "gyro_x","gyro_y","gyro_z",
            "accel_x","accel_y","accel_z",
            "linaccel_x","linaccel_y","linaccel_z",
            "temp_c"
        ]}
        if not self.ok:
            return out
        try:
            cal = self._read(self.REG_CALIB_STAT, 1)[0]
            out["sys_cal"]   = (cal >> 6) & 0x03
            out["gyro_cal"]  = (cal >> 4) & 0x03
            out["accel_cal"] = (cal >> 2) & 0x03
            out["mag_cal"]   =  cal       & 0x03

            # Euler angles: 1 LSB = 1/16 degree
            euler = self._read(self.REG_EULER_H_L, 6)
            out["euler_h"] = self._read_s16(euler, 0) / 16.0
            out["euler_r"] = self._read_s16(euler, 2) / 16.0
            out["euler_p"] = self._read_s16(euler, 4) / 16.0

            # Gyroscope: 1 LSB = 1/900 rad/s (rad/s mode)
            gyro = self._read(self.REG_GYRO_X_L, 6)
            out["gyro_x"] = self._read_s16(gyro, 0) / 900.0
            out["gyro_y"] = self._read_s16(gyro, 2) / 900.0
            out["gyro_z"] = self._read_s16(gyro, 4) / 900.0

            # Accelerometer: 1 LSB = 0.01 m/s²
            accel = self._read(self.REG_ACCEL_X_L, 6)
            out["accel_x"] = self._read_s16(accel, 0) / 100.0
            out["accel_y"] = self._read_s16(accel, 2) / 100.0
            out["accel_z"] = self._read_s16(accel, 4) / 100.0

            # Linear acceleration: 1 LSB = 0.01 m/s²
            lia = self._read(self.REG_LIA_X_L, 6)
            out["linaccel_x"] = self._read_s16(lia, 0) / 100.0
            out["linaccel_y"] = self._read_s16(lia, 2) / 100.0
            out["linaccel_z"] = self._read_s16(lia, 4) / 100.0

            out["temp_c"] = self._read(self.REG_TEMP, 1)[0]
        except Exception as e:
            log.warning("%s: read error — %s", self.label, e)
        return out


# ─────────────────────────────────────────────────────────────────────────────
# MPL3115A2 Driver
# ─────────────────────────────────────────────────────────────────────────────
class MPL3115A2:
    REG_STATUS    = 0x00
    REG_OUT_P_MSB = 0x01
    REG_OUT_T_MSB = 0x04
    REG_CTRL_REG1 = 0x26
    REG_PT_DATA_CFG = 0x13

    def __init__(self, bus: smbus2.SMBus):
        self.bus = bus
        self.ok  = False
        self._init()

    def _write(self, reg, val):
        self.bus.write_byte_data(MPL3115_ADDR, reg, val)

    def _read(self, reg, n):
        return bytes(self.bus.read_i2c_block_data(MPL3115_ADDR, reg, n))

    def _init(self):
        try:
            who = self._read(0x0C, 1)[0]
            if who != 0xC4:
                log.error("MPL3115A2: unexpected WHO_AM_I 0x%02X", who)
                return
            # Altimeter mode, OSR=128, active
            self._write(self.REG_CTRL_REG1, 0xB9)   # ALT=1, OS=111, SBYB=1
            self._write(self.REG_PT_DATA_CFG, 0x07)  # enable data flags
            time.sleep(0.1)
            self.ok = True
            log.info("MPL3115A2: initialized at 0x%02X", MPL3115_ADDR)
        except Exception as e:
            log.error("MPL3115A2: init failed — %s", e)

    def read(self) -> dict:
        out = {"altitude_m": 0.0, "pressure_pa": 0.0, "temp_c": 0.0}
        if not self.ok:
            return out
        try:
            # Trigger one-shot
            ctrl = self._read(self.REG_CTRL_REG1, 1)[0]
            self._write(self.REG_CTRL_REG1, ctrl | 0x02)
            time.sleep(0.55)   # OSR=128 conversion time

            data = self._read(self.REG_OUT_P_MSB, 5)
            # Altitude: Q16.4 fixed point (20-bit signed MSB, 4 fractional bits)
            raw_alt = (data[0] << 24 | data[1] << 16 | data[2] << 8) >> 12
            if raw_alt > (1 << 19):   # sign extend 20-bit
                raw_alt -= (1 << 20)
            out["altitude_m"] = raw_alt / 16.0

            # Temperature: Q8.4 (12-bit signed)
            raw_t = (data[3] << 8 | data[4]) >> 4
            if raw_t > 2047:
                raw_t -= 4096
            out["temp_c"] = raw_t / 16.0

            # Barometer mode read for pressure (switch mode briefly)
            self._write(self.REG_CTRL_REG1, 0x39)   # ALT=0, OS=111, SBYB=1
            self._write(self.REG_CTRL_REG1, 0x3B)   # OST=1 one-shot
            time.sleep(0.55)
            pdata = self._read(self.REG_OUT_P_MSB, 3)
            raw_p = ((pdata[0] << 16) | (pdata[1] << 8) | pdata[2]) >> 4
            out["pressure_pa"] = raw_p / 4.0
            # Return to altimeter mode
            self._write(self.REG_CTRL_REG1, 0xB9)
        except Exception as e:
            log.warning("MPL3115A2: read error — %s", e)
        return out


# ─────────────────────────────────────────────────────────────────────────────
# TeraRanger Evo 60m Driver (I2C mode)
# ─────────────────────────────────────────────────────────────────────────────
class TeraRangerEvo:
    """
    TeraRanger Evo in I2C mode.
    Send 0x00 to trigger; read 4 bytes back: [high, low, crc, 0x00]
    Distance = (byte0 << 8 | byte1) / 1000.0  meters
    0x0001 = too close  /  0xFFFF = out of range / no target
    """
    TRIGGER_CMD = 0x00

    def __init__(self, bus: smbus2.SMBus):
        self.bus = bus
        self.ok  = False
        self._init()

    def _crc8(self, data: bytes) -> int:
        crc = 0
        for b in data:
            crc ^= b
            for _ in range(8):
                crc = ((crc << 1) ^ 0x31) & 0xFF if crc & 0x80 else (crc << 1) & 0xFF
        return crc

    def _init(self):
        try:
            # Send a trigger to verify the device responds
            self.bus.write_byte(TERARANGER_ADDR, self.TRIGGER_CMD)
            time.sleep(0.01)
            self.bus.read_i2c_block_data(TERARANGER_ADDR, 0, 4)
            self.ok = True
            log.info("TeraRanger Evo: initialized at 0x%02X", TERARANGER_ADDR)
        except Exception as e:
            log.error("TeraRanger Evo: init failed — %s", e)

    def read(self) -> dict:
        out = {"distance_m": 0.0}
        if not self.ok:
            return out
        try:
            self.bus.write_byte(TERARANGER_ADDR, self.TRIGGER_CMD)
            time.sleep(0.005)
            data = bytes(self.bus.read_i2c_block_data(TERARANGER_ADDR, 0, 4))

            expected_crc = self._crc8(data[:2])
            if data[2] != expected_crc:
                log.debug("TeraRanger: CRC mismatch")
                return out

            raw = (data[0] << 8) | data[1]
            if raw in (0x0000, 0xFFFF, 0x0001):   # invalid readings
                out["distance_m"] = float("nan")
            else:
                out["distance_m"] = raw / 1000.0
        except Exception as e:
            log.warning("TeraRanger Evo: read error — %s", e)
        return out


# ─────────────────────────────────────────────────────────────────────────────
# Excel Export
# ─────────────────────────────────────────────────────────────────────────────

# Colour palette
CLR_HEADER_BNO1  = "1F4E79"   # dark blue
CLR_HEADER_BNO2  = "375623"   # dark green
CLR_HEADER_MPL   = "7B2D00"   # dark orange
CLR_HEADER_EVO   = "4B0082"   # dark purple
CLR_HEADER_TIME  = "222222"   # near-black
CLR_SUBHDR       = "D9D9D9"   # light grey
CLR_FONT_LIGHT   = "FFFFFF"

THIN = Side(border_style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Column definitions: (header, field_name_in_SensorRecord, number_format, width)
COLUMNS = [
    # Timestamp
    ("Time (s)",         "timestamp_s",     "0.000",  10),

    # ── BNO055 #1 ────────────────────────────────────────────────────────────
    ("BNO1 OK",          "bno1_ok",         "General", 7),
    ("SYS Cal",          "bno1_sys_cal",    "0",       6),
    ("GYR Cal",          "bno1_gyro_cal",   "0",       6),
    ("ACC Cal",          "bno1_accel_cal",  "0",       6),
    ("MAG Cal",          "bno1_mag_cal",    "0",       6),
    ("Euler H (°)",      "bno1_euler_h",    "0.000",  10),
    ("Euler R (°)",      "bno1_euler_r",    "0.000",  10),
    ("Euler P (°)",      "bno1_euler_p",    "0.000",  10),
    ("Gyro X (rad/s)",   "bno1_gyro_x",     "0.00000",12),
    ("Gyro Y (rad/s)",   "bno1_gyro_y",     "0.00000",12),
    ("Gyro Z (rad/s)",   "bno1_gyro_z",     "0.00000",12),
    ("Accel X (m/s²)",   "bno1_accel_x",    "0.000",  11),
    ("Accel Y (m/s²)",   "bno1_accel_y",    "0.000",  11),
    ("Accel Z (m/s²)",   "bno1_accel_z",    "0.000",  11),
    ("LinAcc X (m/s²)",  "bno1_linaccel_x", "0.000",  12),
    ("LinAcc Y (m/s²)",  "bno1_linaccel_y", "0.000",  12),
    ("LinAcc Z (m/s²)",  "bno1_linaccel_z", "0.000",  12),
    ("Temp (°C)",        "bno1_temp_c",     "0.0",     8),

    # ── BNO055 #2 ────────────────────────────────────────────────────────────
    ("BNO2 OK",          "bno2_ok",         "General", 7),
    ("SYS Cal",          "bno2_sys_cal",    "0",       6),
    ("GYR Cal",          "bno2_gyro_cal",   "0",       6),
    ("ACC Cal",          "bno2_accel_cal",  "0",       6),
    ("MAG Cal",          "bno2_mag_cal",    "0",       6),
    ("Euler H (°)",      "bno2_euler_h",    "0.000",  10),
    ("Euler R (°)",      "bno2_euler_r",    "0.000",  10),
    ("Euler P (°)",      "bno2_euler_p",    "0.000",  10),
    ("Gyro X (rad/s)",   "bno2_gyro_x",     "0.00000",12),
    ("Gyro Y (rad/s)",   "bno2_gyro_y",     "0.00000",12),
    ("Gyro Z (rad/s)",   "bno2_gyro_z",     "0.00000",12),
    ("Accel X (m/s²)",   "bno2_accel_x",    "0.000",  11),
    ("Accel Y (m/s²)",   "bno2_accel_y",    "0.000",  11),
    ("Accel Z (m/s²)",   "bno2_accel_z",    "0.000",  11),
    ("LinAcc X (m/s²)",  "bno2_linaccel_x", "0.000",  12),
    ("LinAcc Y (m/s²)",  "bno2_linaccel_y", "0.000",  12),
    ("LinAcc Z (m/s²)",  "bno2_linaccel_z", "0.000",  12),
    ("Temp (°C)",        "bno2_temp_c",     "0.0",     8),

    # ── MPL3115A2 ─────────────────────────────────────────────────────────────
    ("MPL OK",           "mpl_ok",          "General", 7),
    ("Altitude (m)",     "mpl_altitude_m",  "0.000",  11),
    ("Pressure (Pa)",    "mpl_pressure_pa", "0.0",    12),
    ("Temp (°C)",        "mpl_temp_c",      "0.00",    8),

    # ── TeraRanger Evo ────────────────────────────────────────────────────────
    ("EVO OK",           "evo_ok",          "General", 7),
    ("Distance (m)",     "evo_distance_m",  "0.000",  11),
]

# Group spans for the top merged header row: (label, color, start_col, end_col)
def _build_group_spans():
    groups = [("Timestamp", CLR_HEADER_TIME, 1, 1)]
    bno1_start = 2
    bno1_end   = bno1_start + 17   # 18 BNO1 columns
    groups.append(("BNO055 #1  (0x28)", CLR_HEADER_BNO1, bno1_start, bno1_end))
    bno2_start = bno1_end + 1
    bno2_end   = bno2_start + 17
    groups.append(("BNO055 #2  (0x29)", CLR_HEADER_BNO2, bno2_start, bno2_end))
    mpl_start  = bno2_end + 1
    mpl_end    = mpl_start + 3     # 4 MPL columns
    groups.append(("MPL3115A2  (0x60)", CLR_HEADER_MPL,  mpl_start,  mpl_end))
    evo_start  = mpl_end + 1
    evo_end    = evo_start + 1     # 2 EVO columns
    groups.append(("TeraRanger Evo 60m  (0x31)", CLR_HEADER_EVO, evo_start, evo_end))
    return groups


def _cell_color(col_idx: int, row: int) -> Optional[str]:
    """Alternating row tint per sensor group (very subtle)."""
    return None   # keep data rows white for readability


def export_xlsx(records: list[SensorRecord], filename: str):
    log.info("Writing %d records to %s …", len(records), filename)
    wb = Workbook()

    # ── Raw Data sheet ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Raw Data"
    ws.freeze_panes = "B3"   # freeze timestamp col + both header rows

    groups = _build_group_spans()

    # Row 1 — Merged sensor group headers
    for label, color, c_start, c_end in groups:
        cell = ws.cell(row=1, column=c_start, value=label)
        cell.font      = Font(name="Arial", bold=True, size=11, color=CLR_FONT_LIGHT)
        cell.fill      = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border    = BORDER
        if c_start != c_end:
            ws.merge_cells(
                start_row=1, start_column=c_start,
                end_row=1,   end_column=c_end
            )

    ws.row_dimensions[1].height = 22

    # Row 2 — Per-column sub-headers
    for c_idx, (hdr, _, fmt, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=c_idx, value=hdr)
        cell.font      = Font(name="Arial", bold=True, size=9)
        cell.fill      = PatternFill("solid", fgColor=CLR_SUBHDR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    ws.row_dimensions[2].height = 30

    # Data rows
    field_names = [col[1] for col in COLUMNS]
    fmt_map     = {col[1]: col[2] for col in COLUMNS}

    for r_idx, rec in enumerate(records, start=3):
        for c_idx, fname in enumerate(field_names, start=1):
            val  = getattr(rec, fname)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="right" if isinstance(val, float) else "center")
            cell.border    = BORDER
            cell.number_format = fmt_map[fname]

    # ── Statistics sheet ───────────────────────────────────────────────────────
    ws_stats = wb.create_sheet("Statistics")
    ws_stats.column_dimensions["A"].width = 22
    ws_stats.column_dimensions["B"].width = 14
    ws_stats.column_dimensions["C"].width = 14
    ws_stats.column_dimensions["D"].width = 14
    ws_stats.column_dimensions["E"].width = 14
    ws_stats.column_dimensions["F"].width = 14

    stat_headers = ["Measurement", "Mean", "Std Dev", "Min", "Max", "Samples"]
    for c, h in enumerate(stat_headers, 1):
        cell = ws_stats.cell(row=1, column=c, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color=CLR_FONT_LIGHT)
        cell.fill      = PatternFill("solid", fgColor=CLR_HEADER_TIME)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = BORDER

    # Numeric sensor columns for statistics
    stat_cols = [
        ("BNO1 Euler H",    "bno1_euler_h"),
        ("BNO1 Euler R",    "bno1_euler_r"),
        ("BNO1 Euler P",    "bno1_euler_p"),
        ("BNO1 Gyro X",     "bno1_gyro_x"),
        ("BNO1 Gyro Y",     "bno1_gyro_y"),
        ("BNO1 Gyro Z",     "bno1_gyro_z"),
        ("BNO1 Accel X",    "bno1_accel_x"),
        ("BNO1 Accel Y",    "bno1_accel_y"),
        ("BNO1 Accel Z",    "bno1_accel_z"),
        ("BNO1 LinAcc X",   "bno1_linaccel_x"),
        ("BNO1 LinAcc Y",   "bno1_linaccel_y"),
        ("BNO1 LinAcc Z",   "bno1_linaccel_z"),
        ("BNO2 Euler H",    "bno2_euler_h"),
        ("BNO2 Euler R",    "bno2_euler_r"),
        ("BNO2 Euler P",    "bno2_euler_p"),
        ("BNO2 Gyro X",     "bno2_gyro_x"),
        ("BNO2 Gyro Y",     "bno2_gyro_y"),
        ("BNO2 Gyro Z",     "bno2_gyro_z"),
        ("BNO2 Accel X",    "bno2_accel_x"),
        ("BNO2 Accel Y",    "bno2_accel_y"),
        ("BNO2 Accel Z",    "bno2_accel_z"),
        ("BNO2 LinAcc X",   "bno2_linaccel_x"),
        ("BNO2 LinAcc Y",   "bno2_linaccel_y"),
        ("BNO2 LinAcc Z",   "bno2_linaccel_z"),
        ("BNO1-BNO2 ΔH",   None),   # computed below
        ("BNO1-BNO2 ΔR",   None),
        ("BNO1-BNO2 ΔP",   None),
        ("MPL Altitude",    "mpl_altitude_m"),
        ("MPL Pressure",    "mpl_pressure_pa"),
        ("MPL Temp",        "mpl_temp_c"),
        ("EVO Distance",    "evo_distance_m"),
    ]

    # Map field → excel column letter in Raw Data sheet (1-indexed)
    raw_col_map = {fname: get_column_letter(i + 1) for i, (_, fname, _, _) in enumerate(COLUMNS)}

    n = len(records)
    data_start = 3
    data_end   = data_start + n - 1

    for r, (label, fname) in enumerate(stat_cols, start=2):
        ws_stats.cell(row=r, column=1, value=label).font = Font(name="Arial", size=9)
        ws_stats.cell(row=r, column=1).border = BORDER

        if fname is not None:
            col_letter = raw_col_map[fname]
            ref = f"'Raw Data'!{col_letter}{data_start}:{col_letter}{data_end}"
        else:
            # BNO delta: recompute inline — build column references for H, R, or P
            axis = label[-1]   # H, R, or P
            axis_map = {"H": ("bno1_euler_h", "bno2_euler_h"),
                        "R": ("bno1_euler_r", "bno2_euler_r"),
                        "P": ("bno1_euler_p", "bno2_euler_p")}
            col1 = raw_col_map[axis_map[axis][0]]
            col2 = raw_col_map[axis_map[axis][1]]
            # Delta stats require a helper column; leave formula note instead
            for c in range(2, 7):
                ws_stats.cell(row=r, column=c, value="→ see 'Delta' sheet").font = Font(
                    name="Arial", size=9, italic=True, color="888888")
            continue

        mean_f  = f"=AVERAGE({ref})"
        std_f   = f"=STDEV({ref})"
        min_f   = f"=MIN({ref})"
        max_f   = f"=MAX({ref})"
        count_f = f"=COUNT({ref})"

        for c, formula in enumerate([mean_f, std_f, min_f, max_f, count_f], start=2):
            cell = ws_stats.cell(row=r, column=c, value=formula)
            cell.font         = Font(name="Arial", size=9)
            cell.number_format = "0.00000"
            cell.border        = BORDER

    ws_stats.row_dimensions[1].height = 18
    ws_stats.freeze_panes = "B2"

    # ── Delta (BNO1 vs BNO2) sheet ────────────────────────────────────────────
    ws_delta = wb.create_sheet("BNO Delta")
    delta_headers = ["Time (s)", "ΔEuler H (°)", "ΔEuler R (°)", "ΔEuler P (°)",
                     "ΔGyro X", "ΔGyro Y", "ΔGyro Z",
                     "ΔAccel X", "ΔAccel Y", "ΔAccel Z"]
    delta_widths = [10, 12, 12, 12, 12, 12, 12, 12, 12, 12]

    delta_field_pairs = [
        ("bno1_euler_h", "bno2_euler_h"),
        ("bno1_euler_r", "bno2_euler_r"),
        ("bno1_euler_p", "bno2_euler_p"),
        ("bno1_gyro_x",  "bno2_gyro_x"),
        ("bno1_gyro_y",  "bno2_gyro_y"),
        ("bno1_gyro_z",  "bno2_gyro_z"),
        ("bno1_accel_x", "bno2_accel_x"),
        ("bno1_accel_y", "bno2_accel_y"),
        ("bno1_accel_z", "bno2_accel_z"),
    ]

    for c, (h, w) in enumerate(zip(delta_headers, delta_widths), 1):
        cell = ws_delta.cell(row=1, column=c, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color=CLR_FONT_LIGHT)
        cell.fill      = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")
        cell.border    = BORDER
        ws_delta.column_dimensions[get_column_letter(c)].width = w

    time_col   = raw_col_map["timestamp_s"]
    for row_offset in range(n):
        raw_row = data_start + row_offset
        out_row = row_offset + 2
        # Time
        cell = ws_delta.cell(row=out_row, column=1,
                             value=f"='Raw Data'!{time_col}{raw_row}")
        cell.number_format = "0.000"
        cell.font = Font(name="Arial", size=9)
        cell.border = BORDER

        for c_idx, (f1, f2) in enumerate(delta_field_pairs, start=2):
            col1 = raw_col_map[f1]
            col2 = raw_col_map[f2]
            cell = ws_delta.cell(row=out_row, column=c_idx,
                                 value=f"='Raw Data'!{col1}{raw_row}-'Raw Data'!{col2}{raw_row}")
            cell.number_format = "0.00000"
            cell.font   = Font(name="Arial", size=9)
            cell.border = BORDER

    ws_delta.freeze_panes = "B2"

    # ── Metadata sheet ────────────────────────────────────────────────────────
    ws_meta = wb.create_sheet("Metadata")
    meta = [
        ("Run Date",        datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Duration (s)",    LOG_DURATION_S),
        ("Target Rate (Hz)",SAMPLE_HZ),
        ("Total Samples",   len(records)),
        ("Actual Rate (Hz)",f"={len(records)}/{LOG_DURATION_S:.1f}" if LOG_DURATION_S else "N/A"),
        ("BNO055 #1 Addr",  "0x28  (ADDR pin LOW)"),
        ("BNO055 #2 Addr",  "0x29  (ADDR pin HIGH)"),
        ("MPL3115A2 Addr",  "0x60"),
        ("TeraRanger Addr", "0x31"),
        ("I2C Bus",         f"bus {I2C_BUS}  →  Pin 3 SDA / Pin 5 SCL"),
        ("Purpose",         "Sensor error / drift characterisation for Kalman filter design"),
    ]
    ws_meta.column_dimensions["A"].width = 22
    ws_meta.column_dimensions["B"].width = 40
    for r, (k, v) in enumerate(meta, 1):
        ka = ws_meta.cell(row=r, column=1, value=k)
        ka.font   = Font(name="Arial", bold=True, size=10)
        ka.border = BORDER
        vb = ws_meta.cell(row=r, column=2, value=v)
        vb.font   = Font(name="Arial", size=10)
        vb.border = BORDER

    wb.save(filename)
    log.info("Saved: %s", filename)


# ─────────────────────────────────────────────────────────────────────────────
# Main Loop
# ─────────────────────────────────────────────────────────────────────────────
def main():
    log.info("Opening I2C bus %d …", I2C_BUS)
    bus = smbus2.SMBus(I2C_BUS)
    time.sleep(0.1)

    bno1 = BNO055(bus, BNO055_ADDR_A, "BNO055-A")
    bno2 = BNO055(bus, BNO055_ADDR_B, "BNO055-B")
    mpl  = MPL3115A2(bus)
    evo  = TeraRangerEvo(bus)

    records:  list[SensorRecord] = []
    t_start = time.monotonic()
    t_end   = t_start + LOG_DURATION_S

    log.info("Logging for %.0f minutes at %d Hz …", LOG_DURATION_S / 60, SAMPLE_HZ)

    try:
        while True:
            t_now = time.monotonic()
            if t_now >= t_end:
                break

            loop_start = t_now
            elapsed    = t_now - t_start

            rec = SensorRecord(timestamp_s=round(elapsed, 4))

            # BNO055 #1
            b1 = bno1.read()
            rec.bno1_ok        = bno1.ok
            rec.bno1_sys_cal   = b1["sys_cal"]
            rec.bno1_gyro_cal  = b1["gyro_cal"]
            rec.bno1_accel_cal = b1["accel_cal"]
            rec.bno1_mag_cal   = b1["mag_cal"]
            rec.bno1_euler_h   = b1["euler_h"]
            rec.bno1_euler_r   = b1["euler_r"]
            rec.bno1_euler_p   = b1["euler_p"]
            rec.bno1_gyro_x    = b1["gyro_x"]
            rec.bno1_gyro_y    = b1["gyro_y"]
            rec.bno1_gyro_z    = b1["gyro_z"]
            rec.bno1_accel_x   = b1["accel_x"]
            rec.bno1_accel_y   = b1["accel_y"]
            rec.bno1_accel_z   = b1["accel_z"]
            rec.bno1_linaccel_x= b1["linaccel_x"]
            rec.bno1_linaccel_y= b1["linaccel_y"]
            rec.bno1_linaccel_z= b1["linaccel_z"]
            rec.bno1_temp_c    = b1["temp_c"]

            # BNO055 #2
            b2 = bno2.read()
            rec.bno2_ok        = bno2.ok
            rec.bno2_sys_cal   = b2["sys_cal"]
            rec.bno2_gyro_cal  = b2["gyro_cal"]
            rec.bno2_accel_cal = b2["accel_cal"]
            rec.bno2_mag_cal   = b2["mag_cal"]
            rec.bno2_euler_h   = b2["euler_h"]
            rec.bno2_euler_r   = b2["euler_r"]
            rec.bno2_euler_p   = b2["euler_p"]
            rec.bno2_gyro_x    = b2["gyro_x"]
            rec.bno2_gyro_y    = b2["gyro_y"]
            rec.bno2_gyro_z    = b2["gyro_z"]
            rec.bno2_accel_x   = b2["accel_x"]
            rec.bno2_accel_y   = b2["accel_y"]
            rec.bno2_accel_z   = b2["accel_z"]
            rec.bno2_linaccel_x= b2["linaccel_x"]
            rec.bno2_linaccel_y= b2["linaccel_y"]
            rec.bno2_linaccel_z= b2["linaccel_z"]
            rec.bno2_temp_c    = b2["temp_c"]

            # MPL3115A2
            mp = mpl.read()
            rec.mpl_ok         = mpl.ok
            rec.mpl_altitude_m = mp["altitude_m"]
            rec.mpl_pressure_pa= mp["pressure_pa"]
            rec.mpl_temp_c     = mp["temp_c"]

            # TeraRanger Evo
            ev = evo.read()
            rec.evo_ok         = evo.ok
            rec.evo_distance_m = ev["distance_m"]

            records.append(rec)

            # Progress heartbeat every 60 s
            if len(records) % (SAMPLE_HZ * 60) == 0:
                mins_elapsed = elapsed / 60
                log.info("  %.1f / %.1f min  |  %d samples collected",
                         mins_elapsed, LOG_DURATION_S / 60, len(records))

            # Pace to target sample rate
            elapsed_loop = time.monotonic() - loop_start
            sleep_for    = SAMPLE_PERIOD - elapsed_loop
            if sleep_for > 0:
                time.sleep(sleep_for)

    except KeyboardInterrupt:
        log.info("Interrupted by user — saving collected data …")
    finally:
        bus.close()

    if records:
        export_xlsx(records, OUTPUT_FILENAME)
    else:
        log.warning("No records collected — nothing to save.")


if __name__ == "__main__":
    main()